





import streamlit as st
import pandas as pd
import pdfplumber
import re
from difflib import get_close_matches
from io import BytesIO
import os
# ---------------- Your existing extraction functions ----------------
def extract_sb_data(pdf_path):
    sb_data = []

    # Patterns
    sb_regex = r'\b\d{5,8}\b'                    
    date_regex = r'\b\d{2}-[A-Z]{3}-\d{2}\b'    
    iec_regex = r'IEC/Br\s*[:\-]?\s*([A-Z0-9]+)'
    gstin_regex = r'GSTIN/TYPE\s*[:\-]?\s*([A-Z0-9]+)'
    cbcode_regex = r'CB CODE\s*[:\-]?\s*([A-Z0-9]+)'

    # Country list for fuzzy matching
    country_list = [
        "INDIA", "SWEDEN", "GERMANY", "USA", "UNITED STATES", "FRANCE", "ITALY", 
        "CHINA", "JAPAN", "SOUTH KOREA", "THAILAND", "SINGAPORE", "UAE", "BRAZIL",
        "UK", "UNITED KINGDOM", "NORWAY", "FINLAND", "DENMARK", "NETHERLANDS",
        "POLAND", "SPAIN", "CANADA", "AUSTRALIA", "SWITZERLAND", "BELGIUM"
    ]

    with pdfplumber.open(pdf_path) as pdf:
        page = pdf.pages[0]
        text = page.extract_text() or ""

        # --- Extract IEC, GSTIN, CB CODE ---
        iec_value = re.search(iec_regex, text)
        iec_value = iec_value.group(1) if iec_value else ""

        gstin_value = re.search(gstin_regex, text)
        gstin_value = gstin_value.group(1) if gstin_value else ""

        cbcode_value = re.search(cbcode_regex, text)
        cbcode_value = cbcode_value.group(1) if cbcode_value else ""

        # --- Extract FINAL DESTINATION robustly ---
        final_dest_value = ""
        lines = text.split('\n')

        for i, line in enumerate(lines):
            if re.search(r'13\.*\s*COUNTRY\s*OF\s*FINALDESTINATIO', line, re.IGNORECASE):
                after = line.split("13.COUNTRY OF FINALDESTINATIO")[-1].strip()
                candidates = []

                if after:
                    candidates.append(after)

                for next_line in lines[i+1:i+5]:
                    clean_next = next_line.strip()
                    if clean_next:
                        candidates.append(clean_next)

                for cand in candidates:
                    cand_clean = re.sub(r'[^A-Z\s]', '', cand.upper())
                    cand_clean = cand_clean.strip()
                    if not cand_clean:
                        continue

                    match = get_close_matches(cand_clean, country_list, n=1, cutoff=0.5)
                    if match:
                        final_dest_value = match[0]
                        break

                if not final_dest_value and candidates:
                    final_dest_value = candidates[0].strip()
                break

        # --- Extract SB Data Section ---
        for i, line in enumerate(lines):
            if "Port Code SB No SB Date" in line:
                if i + 1 < len(lines):
                    next_line = lines[i + 1]
                    sb_numbers = re.findall(sb_regex, next_line)
                    dates = re.findall(date_regex, next_line)

                    # Port Code (before SB No)
                    port_code = ""
                    if sb_numbers:
                        sb_index = next_line.find(sb_numbers[0])
                        port_code_candidate = next_line[:sb_index].strip()
                        port_code = port_code_candidate.split()[-1] if port_code_candidate else ""

                    # Combine into rows
                    for j in range(max(len(sb_numbers), len(dates))):
                        sb_data.append({
                            "PORT CODE(FROM)": port_code,
                            "SHIPPINGBILL NO": sb_numbers[j] if j < len(sb_numbers) else "",
                            "SHIPPING BILL DATE": dates[j] if j < len(dates) else "",
                            "IE CODE": iec_value,
                            "GSTIN/TYPE": gstin_value,
                            "CB CODE": cbcode_value,
                            "FINAL DESTINATION": final_dest_value,
                            "INVOICE NO": ""   # Placeholder (will be filled later)
                        })

    sb_df = pd.DataFrame(sb_data) if sb_data else None
    return sb_df


def extract_invoice_tables(pdf_path):
    """
    Extract tables from all pages that contain "PART - II - INVOICE DETAILS",
    and also include the first page even if it doesn't contain that text.
    Returns a dictionary: {sheet_name: DataFrame}
    """
    page_tables_dict = {}

    with pdfplumber.open(pdf_path) as pdf:
        for i, page in enumerate(pdf.pages, start=1):
            text = page.extract_text()
            tables = page.extract_tables()

            if (text and "PART - II - INVOICE DETAILS" in text) or i == 1:
                if tables:
                    page_df = pd.concat([pd.DataFrame(tbl) for tbl in tables], ignore_index=True)
                    sheet_name = f"Page_{i}"
                    page_tables_dict[sheet_name] = page_df
                else:
                    page_tables_dict[f"Page_{i}"] = pd.DataFrame([["No table found on this page"]])

    return page_tables_dict

def extract_invoice_details_from_all_pages(tables_dict, sb_df=None):
    """
    Loop over all pages after Page_1, check if J13 contains
    "2.BUYER'S NAME & ADDRESS", and extract invoice/drawee/goods details.
    Append each page's data as a new row to sb_df.
    """
    if sb_df is None:
        sb_df = pd.DataFrame()

    for page_name, page_df in tables_dict.items():
        # Skip Page_1
        if page_name == "Page_1":
            continue

        page_df = page_df.fillna("").astype(str)
        try:
            # Check if row 13, col J (index 12, 9) contains "2.BUYER'S NAME & ADDRESS"
            if page_df.shape[0] >= 13 and page_df.shape[1] >= 10:
                check_cell = page_df.iat[12, 9].strip().upper()
                if "2.BUYER'S NAME & ADDRESS".upper() in check_cell:
                    # ✅ Extract invoice no & date from C12
                    invoice_no, invoice_date = "", ""
                    if page_df.shape[0] >= 12 and page_df.shape[1] >= 3:
                        cell_val = page_df.iat[11, 2].strip()
                        match = re.match(r"([A-Za-z0-9/\\-]+)\s+(\d{2}/\d{2}/\d{4})", cell_val)
                        if match:
                            invoice_no = match.group(1)
                            invoice_date = match.group(2)
                        else:
                            invoice_no = cell_val

                    # ✅ Drawee Name from J14
                    drawee_name = page_df.iat[13, 9].strip() if page_df.shape[0] >= 14 else ""

                    # ✅ Drawee Address from J15–J18
                    drawee_address_parts = []
                    for r in range(14, min(18, page_df.shape[0])):
                        val = page_df.iat[r, 9].strip()
                        if val and len(val) > 2:  # filter out stray characters
                            drawee_address_parts.append(val)
                    drawee_address = " ".join(drawee_address_parts)
                    

                    # ✅ Goods Description from E29
                    goods_desc = page_df.iat[28, 4].strip() if page_df.shape[0] >= 29 else ""

                    # ✅ PORT OF DESTINATION (if exists in Page_1)
                    port_of_dest = ""
                    if "Page_1" in tables_dict and tables_dict["Page_1"].shape[0] >= 14 and tables_dict["Page_1"].shape[1] >= 30:
                        port_of_dest = tables_dict["Page_1"].iat[13, 29].strip()

                    # --- Create new row dict ---
                    new_row = {
                        "INVOICE NO": invoice_no,
                        "INVOICE DATE": invoice_date,
                        "DRAWEE NAME": drawee_name,
                        "DRAWEE ADDRESS": drawee_address,
                        "GOODS DESCRIPTION": goods_desc,
                        "PORT OF DESTINATION": port_of_dest
                    }

                    # Append to SB DataFrame
                    sb_df = pd.concat([sb_df, pd.DataFrame([new_row])], ignore_index=True)

        except Exception as e:
            print(f"Error processing {page_name}: {e}")

    return sb_df

def get_port_of_destination(tables_dict):
    """
    Extract PORT OF DESTINATION from Page_1 cell AD14
    """
    port_of_dest_value = ""

    if "Page_1" in tables_dict:
        page1_df = tables_dict["Page_1"].fillna("").astype(str)
        try:
            # Check if row 14 and column AD exist
            if page1_df.shape[0] >= 14 and page1_df.shape[1] >= 30:
                port_of_dest_value = page1_df.iat[13, 29].strip()
                print(f"🔍 Port of Destination (AD14): '{port_of_dest_value}'")
        except Exception as e:
            print(f"Error extracting Port of Destination from Page_1: {e}")
    else:
        print("⚠️ Page_1 not found in tables_dict.")

    return port_of_dest_value


def save_sb_and_tables(sb_df, tables_dict, sb_output_path, tables_output_path):
    os.makedirs(os.path.dirname(sb_output_path), exist_ok=True)
    os.makedirs(os.path.dirname(tables_output_path), exist_ok=True)

    if sb_df is not None and not sb_df.empty:
        sb_df.to_excel(sb_output_path, index=False)
        print(f"SB Data saved to: {sb_output_path}")
    else:
        print("No SB Data found to save.")

    if tables_dict:
        with pd.ExcelWriter(tables_output_path, engine='openpyxl') as writer:
            for sheet_name, df in tables_dict.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        print(f"Invoice Tables saved to: {tables_output_path}")
    else:
        print("No Invoice Tables found to save.")

# ---------------- Streamlit App ----------------
st.set_page_config(page_title="Multi-PDF SB Data Extractor", layout="wide")
st.title("📄 Multi-PDF SB Data Extractor")

# File uploader (multiple PDFs)
uploaded_files = st.file_uploader(
    "Upload PDF files (multiple allowed)", type=["pdf"], accept_multiple_files=True
)

if uploaded_files:
    st.info("Processing PDFs... This may take a few seconds.")
    combined_sb_df = pd.DataFrame()
    
    for uploaded_file in uploaded_files:
        # Save uploaded file temporarily
        pdf_path = f"temp_uploaded.pdf"
        with open(pdf_path, "wb") as f:
            f.write(uploaded_file.read())
        
        # Extract SB Data
        sb_df = extract_sb_data(pdf_path)
        
        # Extract invoice tables to get invoice/buyer info
        invoice_tables_dict = extract_invoice_tables(pdf_path)
        sb_df = extract_invoice_details_from_all_pages(invoice_tables_dict, sb_df=sb_df)
        
        if sb_df is not None and not sb_df.empty:
            combined_sb_df = pd.concat([combined_sb_df, sb_df], ignore_index=True)
    
    if not combined_sb_df.empty:
        # Fill SB-level columns
        sb_columns_to_fill = [
            "PORT CODE(FROM)", "SHIPPINGBILL NO", "SHIPPING BILL DATE", "IE CODE", 
            "GSTIN/TYPE", "CB CODE", "FINAL DESTINATION"
        ]
        combined_sb_df[sb_columns_to_fill] = combined_sb_df[sb_columns_to_fill].ffill()

        # Remove rows where INVOICE NO is blank
        combined_sb_df = combined_sb_df[
            combined_sb_df["INVOICE NO"].notna() & (combined_sb_df["INVOICE NO"] != "")
        ]

        # Clean DRAWEE ADDRESS
        if "DRAWEE ADDRESS" in combined_sb_df.columns:
            combined_sb_df["DRAWEE ADDRESS"] = combined_sb_df["DRAWEE ADDRESS"].str.replace(r'Y\s*\n', '', regex=True).str.strip()
        
        # Display combined SB Data
        st.subheader("📊 Combined SB Data")
        st.dataframe(combined_sb_df)
        
        # Download combined SB Data as Excel
        towrite = BytesIO()
        combined_sb_df.to_excel(towrite, index=False, engine='openpyxl')
        towrite.seek(0)
        st.download_button(
            label="⬇️ Download Combined SB Data as Excel",
            data=towrite,
            file_name="Combined_SB_Data.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    else:
        st.warning("No SB Data found in the uploaded PDFs.")


st.subheader("🔍 Filter by Shipping Bill No. (Multiple Allowed)")

# Input for SB No. (comma-separated)
sb_input = st.text_input("Enter Shipping Bill Numbers (comma-separated):")

if sb_input:
    # Split by comma and strip spaces
    sb_list = [sb.strip() for sb in sb_input.split(" ") if sb.strip()]
    
    # Filter the DataFrame
    filtered_df = combined_sb_df[combined_sb_df["SHIPPINGBILL NO"].astype(str).isin(sb_list)]
    
    if not filtered_df.empty:
        st.write(f"Showing data for SB No.: {', '.join(sb_list)}")
        st.dataframe(filtered_df)
        
        # Download filtered SB Data as Excel
        towrite_filtered = BytesIO()
        filtered_df.to_excel(towrite_filtered, index=False, engine='openpyxl')
        towrite_filtered.seek(0)
        
        st.download_button(
            label=f"⬇️ Download Filtered SB Data",
            data=towrite_filtered,
            file_name=f"Filtered_SB_Data.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    else:
        st.warning(f"No data found for SB No.: {', '.join(sb_list)}")


st.subheader("📥 Upload Excel with Shipping Bill Numbers to Fill Data")

uploaded_excel = st.file_uploader(
    "Upload Excel file containing 'SHIPPINGBILL NO' column", type=["xlsx"]
)

import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook

st.subheader("📥 Upload Excel with Shipping Bill Numbers to Fill Data")

# --- Step 1: Choose data source for combined_sb_df ---
st.write("### Step 1: Choose Data Source for Shipping Bill Details")

data_source = st.radio(
    "Select data source for filling details:",
    ("Use existing combined_sb_df", "Upload another Excel file")
)

# Initialize effective_combined_df
effective_combined_df = None

if data_source == "Use existing combined_sb_df":
    if 'combined_sb_df' not in locals() or combined_sb_df.empty:
        st.error("❌ No existing combined_sb_df found in memory.")
    else:
        effective_combined_df = combined_sb_df
        st.success("✅ Using existing combined_sb_df loaded in memory.")

else:
    uploaded_combined = st.file_uploader(
        "Upload Excel file to use as combined_sb_df",
        type=["xlsx"],
        key="combined_source"
    )
    if uploaded_combined:
        effective_combined_df = pd.read_excel(uploaded_combined)
        st.success("✅ Uploaded Excel loaded as combined_sb_df.")

# --- Step 2: Upload target Excel to fill ---
uploaded_excel = st.file_uploader(
    "Upload Excel file containing 'SHIPPINGBILL NO' column to fill data",
    type=["xlsx"],
    key="target_excel"
)

if uploaded_excel and effective_combined_df is not None and not effective_combined_df.empty:
    # Read uploaded target Excel
    user_sb_df = pd.read_excel(uploaded_excel)
    
    if "SHIPPINGBILL NO" not in user_sb_df.columns:
        st.error("❌ The uploaded Excel must contain a column named 'SHIPPINGBILL NO'.")
    else:
        # Normalize both for comparison
        user_sb_df["SHIPPINGBILL NO"] = user_sb_df["SHIPPINGBILL NO"].astype(str).str.strip()
        effective_combined_df["SHIPPINGBILL NO"] = effective_combined_df["SHIPPINGBILL NO"].astype(str).str.strip()

        # Extract relevant columns
        extracted_cols = [
            "SHIPPINGBILL NO", "PORT CODE(FROM)", "SHIPPING BILL DATE", "IE CODE", "GSTIN/TYPE",
            "CB CODE", "FINAL DESTINATION", "INVOICE NO", "INVOICE DATE", "DRAWEE NAME",
            "DRAWEE ADDRESS", "GOODS DESCRIPTION", "PORT OF DESTINATION"
        ]
        extracted_data = effective_combined_df[
            [c for c in extracted_cols if c in effective_combined_df.columns]
        ].drop_duplicates(subset=["SHIPPINGBILL NO"])

        extracted_dict = extracted_data.set_index("SHIPPINGBILL NO").to_dict(orient="index")

        # Load workbook preserving formatting
        in_memory_file = BytesIO(uploaded_excel.getvalue())
        wb = load_workbook(in_memory_file)
        ws = wb.active

        header_map = {cell.value.strip(): cell.column_letter for cell in ws[1] if cell.value}

        for row in ws.iter_rows(min_row=2):
            sb_cell = None
            for cell in row:
                if cell.column_letter == header_map.get("SHIPPINGBILL NO"):
                    sb_cell = cell
                    break

            if sb_cell and sb_cell.value:
                sb_no = str(sb_cell.value).strip()
                if sb_no in extracted_dict:
                    extracted_row = extracted_dict[sb_no]
                    for field, value in extracted_row.items():
                        if field in header_map:
                            target_cell = ws[f"{header_map[field]}{cell.row}"]
                            if (target_cell.value is None or str(target_cell.value).strip() == "") and value not in [None, ""]:
                                target_cell.value = value

        filled_excel = BytesIO()
        wb.save(filled_excel)
        filled_excel.seek(0)

        st.success("✅ Data from PDFs (or uploaded combined file) has been filled into your uploaded Excel — formatting preserved!")

        st.dataframe(pd.read_excel(filled_excel))

        st.download_button(
            label="⬇️ Download Filled Excel (Format Preserved)",
            data=filled_excel,
            file_name="Filled_SB_Data_Formatted.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
