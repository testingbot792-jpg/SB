# import win32com.client
from pathlib import Path

import os
# import win32com.client
from pathlib import Path

def download_pdfs_from_outlook(folder_name, sb_numbers, download_dir="temp_pdfs"):
    """
    Connects to Outlook, searches the given folder for emails that have PDF attachments
    containing any of the specified Shipping Bill Numbers in their filename.
    Downloads those attachments into download_dir (absolute path).
    """
    # ✅ Use absolute, guaranteed-valid path (under user's Documents)
    base_path = Path.home() / "Documents" / download_dir
    base_path.mkdir(parents=True, exist_ok=True)

    outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
    inbox = outlook.GetDefaultFolder(6)  # 6 = Inbox
    target_folder = None

    # Try to find the subfolder "abc" (case-insensitive)
    for folder in inbox.Folders:
        if folder.Name.lower() == folder_name.lower():
            target_folder = folder
            break

    if not target_folder:
        print(f"❌ Folder '{folder_name}' not found in Outlook Inbox.")
        return []

    print(f"📂 Searching Outlook folder '{folder_name}' for matching PDFs...")

    downloaded_files = []
    messages = target_folder.Items

    for msg in messages:
        attachments = msg.Attachments
        for att in attachments:
            if att.FileName.lower().endswith(".pdf"):
                for sb_no in sb_numbers:
                    if str(sb_no) in att.FileName:
                        # ✅ Safe absolute path
                        safe_filename = att.FileName.replace(":", "_").replace("\\", "_").replace("/", "_")
                        save_path = str(base_path / safe_filename)
                        att.SaveAsFile(save_path)
                        downloaded_files.append(save_path)
                        print(f"✅ Downloaded: {att.FileName}")
                        break  # Move to next attachment

    if not downloaded_files:
        print("⚠️ No matching PDF attachments found.")
    else:
        print(f"📦 {len(downloaded_files)} PDFs downloaded to '{base_path}'.")

    return downloaded_files

import streamlit as st
import pandas as pd
import pdfplumber
import re
from difflib import get_close_matches
from io import BytesIO
import os
from openpyxl import load_workbook
# ---------------- Your existing extraction functions ----------------
def extract_sb_data(pdf_path):
    sb_data = []

    # Patterns
    sb_regex = r'\b\d{5,8}\b'                    
    date_regex = r'\b\d{2}-[A-Z]{3}-\d{2}\b'    
    iec_regex = r'IEC/Br\s*[:\-]?\s*([A-Z0-9]+)'
    gstin_regex = r'GSTIN/TYPE\s*[:\-]?\s*([A-Z0-9]+)'
    cbcode_regex = r'CB CODE\s*[:\-]?\s*([A-Z0-9]+)'

    # Country list for fuzzy matching
    country_list = [
        "INDIA", "SWEDEN", "GERMANY", "USA", "UNITED STATES", "FRANCE", "ITALY", 
        "CHINA", "JAPAN", "SOUTH KOREA", "THAILAND", "SINGAPORE", "UAE", "BRAZIL",
        "UK", "UNITED KINGDOM", "NORWAY", "FINLAND", "DENMARK", "NETHERLANDS",
        "POLAND", "SPAIN", "CANADA", "AUSTRALIA", "SWITZERLAND", "BELGIUM"
    ]

    with pdfplumber.open(pdf_path) as pdf:
        page = pdf.pages[0]
        text = page.extract_text() or ""

        # --- Extract IEC, GSTIN, CB CODE ---
        iec_value = re.search(iec_regex, text)
        iec_value = iec_value.group(1) if iec_value else ""

        gstin_value = re.search(gstin_regex, text)
        gstin_value = gstin_value.group(1) if gstin_value else ""

        cbcode_value = re.search(cbcode_regex, text)
        cbcode_value = cbcode_value.group(1) if cbcode_value else ""

        # --- Extract FINAL DESTINATION robustly ---
        final_dest_value = ""
        lines = text.split('\n')

        for i, line in enumerate(lines):
            if re.search(r'13\.*\s*COUNTRY\s*OF\s*FINALDESTINATIO', line, re.IGNORECASE):
                after = line.split("13.COUNTRY OF FINALDESTINATIO")[-1].strip()
                candidates = []

                if after:
                    candidates.append(after)

                for next_line in lines[i+1:i+5]:
                    clean_next = next_line.strip()
                    if clean_next:
                        candidates.append(clean_next)

                for cand in candidates:
                    cand_clean = re.sub(r'[^A-Z\s]', '', cand.upper())
                    cand_clean = cand_clean.strip()
                    if not cand_clean:
                        continue

                    match = get_close_matches(cand_clean, country_list, n=1, cutoff=0.5)
                    if match:
                        final_dest_value = match[0]
                        break

                if not final_dest_value and candidates:
                    final_dest_value = candidates[0].strip()
                break

        # --- Extract SB Data Section ---
        for i, line in enumerate(lines):
            if "Port Code SB No SB Date" in line:
                if i + 1 < len(lines):
                    next_line = lines[i + 1]
                    sb_numbers = re.findall(sb_regex, next_line)
                    dates = re.findall(date_regex, next_line)

                    # Port Code (before SB No)
                    port_code = ""
                    if sb_numbers:
                        sb_index = next_line.find(sb_numbers[0])
                        port_code_candidate = next_line[:sb_index].strip()
                        port_code = port_code_candidate.split()[-1] if port_code_candidate else ""

                    # Combine into rows
                    for j in range(max(len(sb_numbers), len(dates))):
                        sb_data.append({
                            "PORT CODE(FROM)": port_code,
                            "SHIPPINGBILL NO": sb_numbers[j] if j < len(sb_numbers) else "",
                            "SHIPPING BILL DATE": dates[j] if j < len(dates) else "",
                            "IE CODE": iec_value,
                            "GSTIN/TYPE": gstin_value,
                            "CB CODE": cbcode_value,
                            "FINAL DESTINATION": final_dest_value,
                            "INVOICE NO": ""   # Placeholder (will be filled later)
                        })

    sb_df = pd.DataFrame(sb_data) if sb_data else None
    return sb_df


def extract_invoice_tables(pdf_path):
    """
    Extract tables from all pages that contain "PART - II - INVOICE DETAILS",
    and also include the first page even if it doesn't contain that text.
    Returns a dictionary: {sheet_name: DataFrame}
    """
    page_tables_dict = {}

    with pdfplumber.open(pdf_path) as pdf:
        for i, page in enumerate(pdf.pages, start=1):
            text = page.extract_text()
            tables = page.extract_tables()

            if (text and "PART - II - INVOICE DETAILS" in text) or i == 1:
                if tables:
                    page_df = pd.concat([pd.DataFrame(tbl) for tbl in tables], ignore_index=True)
                    sheet_name = f"Page_{i}"
                    page_tables_dict[sheet_name] = page_df
                else:
                    page_tables_dict[f"Page_{i}"] = pd.DataFrame([["No table found on this page"]])

    return page_tables_dict

def extract_invoice_details_from_all_pages(tables_dict, sb_df=None):
    """
    Extracts *all* invoices from all 'PART - II - INVOICE DETAILS' pages,
    and duplicates SB-level info for each invoice.
    """
    if sb_df is None or sb_df.empty:
        sb_df = pd.DataFrame()

    invoice_rows = []

    for page_name, page_df in tables_dict.items():
        if page_name == "Page_1":
            continue

        page_df = page_df.fillna("").astype(str)
        try:
            if page_df.shape[0] >= 13 and page_df.shape[1] >= 10:
                check_cell = page_df.iat[12, 9].strip().upper()
                if "2.BUYER'S NAME & ADDRESS".upper() in check_cell:
                    # Extract invoice number and date
                    invoice_no, invoice_date = "", ""
                    if page_df.shape[0] >= 12 and page_df.shape[1] >= 3:
                        cell_val = page_df.iat[11, 2].strip()
                        match = re.match(r"([A-Za-z0-9/\\-]+)\s+(\d{2}/\d{2}/\d{4})", cell_val)
                        if match:
                            invoice_no = match.group(1)
                            invoice_date = match.group(2)
                        else:
                            invoice_no = cell_val

                    # Extract drawee name
                    drawee_name = page_df.iat[13, 9].strip() if page_df.shape[0] >= 14 else ""

                    # Extract drawee address (next few rows)
                    drawee_address_parts = []
                    for r in range(14, min(19, page_df.shape[0])):
                        val = page_df.iat[r, 9].strip()
                        if len(val) > 2:
                            drawee_address_parts.append(val)
                    drawee_address = " ".join(drawee_address_parts)

                    # Goods description
                    goods_desc = page_df.iat[28, 4].strip() if page_df.shape[0] >= 29 else ""

                    # Port of destination
                    port_of_dest = ""
                    if "Page_1" in tables_dict:
                        page1_df = tables_dict["Page_1"].fillna("").astype(str)
                        if page1_df.shape[0] >= 14 and page1_df.shape[1] >= 30:
                            port_of_dest = page1_df.iat[13, 29].strip()

                    invoice_rows.append({
                        "INVOICE NO": invoice_no,
                        "INVOICE DATE": invoice_date,
                        "DRAWEE NAME": drawee_name,
                        "DRAWEE ADDRESS": drawee_address,
                        "GOODS DESCRIPTION": goods_desc,
                        "PORT OF DESTINATION": port_of_dest
                    })
        except Exception as e:
            print(f"Error processing {page_name}: {e}")

    # ✅ Combine SB data with *all* invoices
    combined_rows = []
    if not sb_df.empty:
        for _, sb_row in sb_df.iterrows():
            for inv in invoice_rows:
                combined_row = {**sb_row.to_dict(), **inv}
                combined_rows.append(combined_row)
    else:
        combined_rows = invoice_rows

    return pd.DataFrame(combined_rows)

def get_port_of_destination(tables_dict):
    """
    Extract PORT OF DESTINATION from Page_1 cell AD14
    """
    port_of_dest_value = ""

    if "Page_1" in tables_dict:
        page1_df = tables_dict["Page_1"].fillna("").astype(str)
        try:
            # Check if row 14 and column AD exist
            if page1_df.shape[0] >= 14 and page1_df.shape[1] >= 30:
                port_of_dest_value = page1_df.iat[13, 29].strip()
                print(f"🔍 Port of Destination (AD14): '{port_of_dest_value}'")
        except Exception as e:
            print(f"Error extracting Port of Destination from Page_1: {e}")
    else:
        print("⚠️ Page_1 not found in tables_dict.")

    return port_of_dest_value


def save_sb_and_tables(sb_df, tables_dict, sb_output_path, tables_output_path):
    os.makedirs(os.path.dirname(sb_output_path), exist_ok=True)
    os.makedirs(os.path.dirname(tables_output_path), exist_ok=True)

    if sb_df is not None and not sb_df.empty:
        sb_df.to_excel(sb_output_path, index=False)
        print(f"SB Data saved to: {sb_output_path}")
    else:
        print("No SB Data found to save.")

    if tables_dict:
        with pd.ExcelWriter(tables_output_path, engine='openpyxl') as writer:
            for sheet_name, df in tables_dict.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        print(f"Invoice Tables saved to: {tables_output_path}")
    else:
        print("No Invoice Tables found to save.")

# ---------------- Streamlit App ----------------
st.set_page_config(page_title="Multi-PDF SB Data Extractor", layout="wide")
st.title("📄 Multi-PDF SB Data Extractor")

# File uploader (multiple PDFs)
uploaded_files = st.file_uploader(
    "Upload PDF files (multiple allowed)", type=["pdf"], accept_multiple_files=True
)

if uploaded_files:
    st.info("Processing PDFs... This may take a few seconds.")
    combined_sb_df = pd.DataFrame()
    
    for uploaded_file in uploaded_files:
        # Save uploaded file temporarily
        pdf_path = f"temp_uploaded.pdf"
        with open(pdf_path, "wb") as f:
            f.write(uploaded_file.read())
        
        # Extract SB Data
        sb_df = extract_sb_data(pdf_path)
        
        # Extract invoice tables to get invoice/buyer info
        invoice_tables_dict = extract_invoice_tables(pdf_path)
        sb_df = extract_invoice_details_from_all_pages(invoice_tables_dict, sb_df=sb_df)
        
        if sb_df is not None and not sb_df.empty:
            combined_sb_df = pd.concat([combined_sb_df, sb_df], ignore_index=True)
    
    if not combined_sb_df.empty:
        # Fill SB-level columns
        sb_columns_to_fill = [
            "PORT CODE(FROM)", "SHIPPINGBILL NO", "SHIPPING BILL DATE", "IE CODE", 
            "GSTIN/TYPE", "CB CODE", "FINAL DESTINATION"
        ]
        combined_sb_df[sb_columns_to_fill] = combined_sb_df[sb_columns_to_fill].ffill()

        # Remove rows where INVOICE NO is blank
        combined_sb_df = combined_sb_df[
            combined_sb_df["INVOICE NO"].notna() & (combined_sb_df["INVOICE NO"] != "")
        ]

        # Display combined SB Data
        st.subheader("📊 Combined SB Data")
        st.dataframe(combined_sb_df)
        
        # Download combined SB Data as Excel
        towrite = BytesIO()
        combined_sb_df.to_excel(towrite, index=False, engine='openpyxl')
        towrite.seek(0)
        st.download_button(
            label="⬇️ Download Combined SB Data as Excel",
            data=towrite,
            file_name="Combined_SB_Data.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    else:
        st.warning("No SB Data found in the uploaded PDFs.")



################################################################################################################################

st.subheader("📥 Upload Excel with Shipping Bill Numbers to Fill Data")


data_source = st.radio(
    "Select data source for filling details:",
    ("Use existing combined_sb_df", "Upload another Excel file")
)

# Initialize effective_combined_df
effective_combined_df = None

if data_source == "Use existing combined_sb_df":
    if 'combined_sb_df' not in locals() or combined_sb_df.empty:
        st.error("❌ No existing file found in memory.")
    else:
        effective_combined_df = combined_sb_df
        st.success("✅ Using existing combined_sb_df loaded in memory.")

else:
    uploaded_combined = st.file_uploader(
        "Upload Excel file to use as combined_sb_df",
        type=["xlsx"],
        key="combined_source"
    )
    if uploaded_combined:
        effective_combined_df = pd.read_excel(uploaded_combined)
        st.success("✅ Uploaded Excel loaded as combined_sb_df.")


# --- Step 2: Upload target Excel (template) ---
uploaded_excel = st.file_uploader(
    "Upload Excel file containing 'SHIPPINGBILL NO' column to fill data",
    type=["xlsx"],
    key="target_excel"
)

if uploaded_excel:
    user_sb_df = pd.read_excel(uploaded_excel)

    if "SHIPPINGBILL NO" not in user_sb_df.columns:
        st.error("❌ The uploaded Excel must contain a column named 'SHIPPINGBILL NO'.")
    else:
        sb_list = user_sb_df["SHIPPINGBILL NO"].astype(str).str.strip().unique().tolist()
        st.info(f"Looking for {len(sb_list)} Shipping Bill PDFs in Outlook folder 'abc'...")

        # 🔹 Auto-download PDFs from Outlook
        downloaded_files = download_pdfs_from_outlook("abc", sb_list)

        if downloaded_files:
            st.success(f"✅ Found and downloaded {len(downloaded_files)} PDF(s). Extracting data...")

            # 🔹 Extract all PDFs into one DataFrame
            extracted_sb_df = pd.DataFrame()
            for pdf_path in downloaded_files:
                sb_df = extract_sb_data(pdf_path)
                invoice_tables_dict = extract_invoice_tables(pdf_path)
                sb_df = extract_invoice_details_from_all_pages(invoice_tables_dict, sb_df=sb_df)
                if sb_df is not None and not sb_df.empty:
                    extracted_sb_df = pd.concat([extracted_sb_df, sb_df], ignore_index=True)

            if not extracted_sb_df.empty:
                st.success("✅ PDF extraction complete. Merging into template...")
                effective_combined_df = extracted_sb_df
            else:
                st.warning("⚠️ No data extracted from the downloaded PDFs.")
        else:
            st.warning("⚠️ No matching PDFs found in Outlook folder 'abc'. Proceeding with existing/combined data if available.")

    # ✅ Only continue if we now have combined data
    if effective_combined_df is not None and not effective_combined_df.empty:
        # Normalize both data sources
        user_sb_df["SHIPPINGBILL NO"] = user_sb_df["SHIPPINGBILL NO"].astype(str).str.strip()
        effective_combined_df["SHIPPINGBILL NO"] = effective_combined_df["SHIPPINGBILL NO"].astype(str).str.strip()

        # Select relevant columns
        extracted_cols = [
            "SHIPPINGBILL NO", "PORT CODE(FROM)", "SHIPPING BILL DATE", "IE CODE", "GSTIN/TYPE",
            "CB CODE", "FINAL DESTINATION", "INVOICE NO", "INVOICE DATE", "DRAWEE NAME",
            "DRAWEE ADDRESS", "GOODS DESCRIPTION", "PORT OF DESTINATION"
        ]
        available_cols = [c for c in extracted_cols if c in effective_combined_df.columns]
        extracted_data = effective_combined_df[available_cols].drop_duplicates(subset=["SHIPPINGBILL NO"])
        extracted_dict = extracted_data.set_index("SHIPPINGBILL NO").to_dict(orient="index")

        st.write(f"🧾 Extracted {len(extracted_data)} unique Shipping Bills from PDFs.")

        # --- Fill the template Excel (preserving formatting)
        in_memory_file = BytesIO(uploaded_excel.getvalue())
        wb = load_workbook(in_memory_file)
        ws = wb.active

        header_map = {str(cell.value).strip(): cell.column_letter for cell in ws[1] if cell.value}

        filled_rows = 0
        # --- Build multi-key lookup using both SB and INVOICE NO ---
        if "INVOICE NO" in effective_combined_df.columns:
            extracted_dict = (
                effective_combined_df
                .set_index(["SHIPPINGBILL NO", "INVOICE NO"])
                .to_dict(orient="index")
            )
        else:
            extracted_dict = (
                effective_combined_df
                .set_index("SHIPPINGBILL NO")
                .to_dict(orient="index")
            )

        filled_rows = 0
        for row in ws.iter_rows(min_row=2):
            sb_value = None
            invoice_value = None

            # Locate SB and Invoice cells
            if "SHIPPINGBILL NO" in header_map:
                sb_cell = ws[f"{header_map['SHIPPINGBILL NO']}{row[0].row}"]
                sb_value = str(sb_cell.value).strip() if sb_cell.value else None
            if "INVOICE NO" in header_map:
                inv_cell = ws[f"{header_map['INVOICE NO']}{row[0].row}"]
                invoice_value = str(inv_cell.value).strip() if inv_cell.value else None

            # Lookup match
            key = None
            if invoice_value and (sb_value, invoice_value) in extracted_dict:
                key = (sb_value, invoice_value)
            elif sb_value in extracted_dict:  # fallback if invoice not found
                key = sb_value

            if key:
                extracted_row = extracted_dict[key]
                for field, value in extracted_row.items():
                    if field in header_map:
                        target_cell = ws[f"{header_map[field]}{row[0].row}"]
                        if (target_cell.value is None or str(target_cell.value).strip() == "") and value not in [None, ""]:
                            target_cell.value = value
                filled_rows += 1

        if filled_rows > 0:
            st.success(f"✅ Filled data for {filled_rows} rows in your Excel template.")
        else:
            st.warning("⚠️ No matching Shipping Bill Numbers found in extracted data.")

        # Save filled Excel
        filled_excel = BytesIO()
        wb.save(filled_excel)
        filled_excel.seek(0)

        # Show preview
        st.dataframe(pd.read_excel(filled_excel))

        st.download_button(
            label="⬇️ Download Filled Excel",
            data=filled_excel,
            file_name="Filled_SB_Data_Formatted.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    else:
        st.error("❌ No extracted data available to fill into the template.")

#######################################################################################################################################
st.subheader("📂 Optional: Upload SIH Excel File to Fill Remittance Data")

uploaded_sih = st.file_uploader(
    "Upload SIH file (Invoice-wise remittance data)",
    type=["xlsx"],
    key="sih_excel"
)

if uploaded_sih and effective_combined_df is not None and not effective_combined_df.empty:
    # --- Read SIH without assuming header ---
    sih_df = pd.read_excel(uploaded_sih, sheet_name="SIH", header=None)

    # --- Find header row in first 5 rows ---
    header_row_index = None
    for i in range(5):
        row_values = sih_df.iloc[i].astype(str).str.strip().tolist()
        if any("Invoice Id" in val for val in row_values):
            header_row_index = i
            break

    if header_row_index is not None:
        # Set proper headers
        sih_df.columns = sih_df.iloc[header_row_index]
        sih_df = sih_df.iloc[header_row_index + 1:].reset_index(drop=True)
        st.success(f"✅ SIH header found at row {header_row_index + 1}")

        # Normalize SIH headers
        sih_df.columns = sih_df.columns.astype(str).str.strip().str.lower()

        # --- Load template Excel ---
        in_memory_file.seek(0)
        wb = load_workbook(in_memory_file)
        ws = wb.active

        # Normalize template headers
        header_map = {str(cell.value).strip().lower(): cell.column_letter for cell in ws[1] if cell.value}

        # --- Mapping SIH columns to template headers ---
        sih_mapping = {
            "invoice id": "invoice no",
            "due date": "due date",
            "usd": "realized amount in remittance currency",
            "amount": "realized amount in invoice currency",
            "drawee name": "drawee name",
            "drawee address": "drawee address"
            # Add more fields here if needed
        }

        # --- Fill extracted PDF data first ---
        for row_idx, row_data in effective_combined_df.iterrows():
            excel_row = row_idx + 2  # Assuming template starts at row 2
            for col_name in effective_combined_df.columns:
                col_key = col_name.strip().lower()
                if col_key in header_map:
                    ws[f"{header_map[col_key]}{excel_row}"].value = row_data[col_name]

        # --- Fill SIH data ---
        filled_sih_rows = 0
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            inv_no_cell_letter = header_map.get("invoice no")
            if not inv_no_cell_letter:
                continue  # Skip if invoice no column not found

            inv_no_cell = ws[f"{inv_no_cell_letter}{row[0].row}"]
            inv_no = str(inv_no_cell.value).strip() if inv_no_cell.value else None

            if inv_no and inv_no in sih_df["invoice id"].values:
                sih_row = sih_df[sih_df["invoice id"] == inv_no].iloc[0]

                for sih_col, template_col in sih_mapping.items():
                    template_col_letter = header_map.get(template_col.lower())
                    if template_col_letter and sih_col in sih_row and pd.notna(sih_row[sih_col]):
                        target_cell = ws[f"{template_col_letter}{row[0].row}"]

                        # Convert due date to datetime if needed
                        if sih_col.lower() == "due date":
                            target_cell.value = pd.to_datetime(sih_row[sih_col])
                        else:
                            target_cell.value = sih_row[sih_col]
                filled_sih_rows += 1

        st.success(f"✅ Filled SIH remittance data for {filled_sih_rows} rows.")

        # --- Save updated Excel ---
        final_excel = BytesIO()
        wb.save(final_excel)
        final_excel.seek(0)
        st.download_button(
            label="⬇️ Download Template with Extracted & SIH Data Filled",
            data=final_excel,
            file_name="Filled_SB_Data_with_SIH.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    else:
        st.warning("⚠️ Could not find 'Invoice Id' in the first 5 rows of SIH file.")

